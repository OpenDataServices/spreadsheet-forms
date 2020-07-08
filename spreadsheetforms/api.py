import datetime
from shutil import copyfile

import openpyxl

from .util import json_append_deep_value, json_get_deep_value, json_set_deep_value


def _is_content_a_guide_field(content):
    return content and content.startswith("SPREADSHEETFORM:")


def _get_guide_field_spec_from_content(content):
    if isinstance(content, str) and content.startswith("SPREADSHEETFORM:SINGLE:"):
        return {"mode": "single", "path": content[23:]}

    if isinstance(content, str) and content.startswith("SPREADSHEETFORM:DOWN:"):
        bits = content.split(":")
        return {
            "mode": "down",
            "list_path": bits[2],
            "item_path": bits[3],
        }

    if isinstance(content, str) and content.startswith("SPREADSHEETFORM:RIGHT:"):
        bits = content.split(":")
        return {
            "mode": "right",
            "list_path": bits[2],
            "item_path": bits[3],
        }

    return None


def _get_cell_value(cell, date_format=None):
    value = cell.value

    if isinstance(value, datetime.datetime) and date_format:
        return value.strftime(date_format)

    return value


def make_empty_form(guide_filename, out_filename):
    copyfile(guide_filename, out_filename)
    workbook = openpyxl.load_workbook(out_filename)

    for worksheet in workbook.worksheets:
        for row_idx in range(1, worksheet.max_row + 1):
            for cell in worksheet[row_idx]:
                if _is_content_a_guide_field(cell.value):
                    worksheet[cell.coordinate] = ""

    workbook.save(out_filename)


def _build_all_configs_in_excel_sheet(worksheet):
    single_configs = {}
    down_configs = {}
    right_configs = {}

    for row_idx in range(1, worksheet.max_row + 1):
        for cell in worksheet[row_idx]:
            guide_field_spec = _get_guide_field_spec_from_content(cell.value)
            if guide_field_spec:
                guide_field_spec["coordinate"] = cell.coordinate
                guide_field_spec["column_letter"] = cell.column_letter
                guide_field_spec["column"] = cell.column
                guide_field_spec["row"] = cell.row
                if guide_field_spec["mode"] == "single":
                    single_configs[guide_field_spec["path"]] = guide_field_spec
                elif guide_field_spec["mode"] == "down":
                    if guide_field_spec["list_path"] in down_configs:
                        down_configs[guide_field_spec["list_path"]].append(
                            guide_field_spec
                        )
                    else:
                        down_configs[guide_field_spec["list_path"]] = [guide_field_spec]
                elif guide_field_spec["mode"] == "right":
                    if guide_field_spec["list_path"] in right_configs:
                        right_configs[guide_field_spec["list_path"]].append(
                            guide_field_spec
                        )
                    else:
                        right_configs[guide_field_spec["list_path"]] = [
                            guide_field_spec
                        ]

    # TODO could check that every down config for each list_path is on the same row. Things will get odd if they are not.

    return single_configs, down_configs, right_configs


def get_data_from_form(guide_filename, in_filename, date_format=None):
    guide_spec = get_guide_spec(guide_filename)
    return get_data_from_form_with_guide_spec(
        guide_spec, in_filename, date_format=date_format
    )


def get_guide_spec(guide_filename):
    guide_workbook = openpyxl.load_workbook(guide_filename, read_only=True)
    guide_spec = {"worksheets": {}}
    for worksheet in guide_workbook.worksheets:
        single_configs, down_configs, right_configs = _build_all_configs_in_excel_sheet(
            worksheet
        )
        guide_spec["worksheets"][worksheet.title] = {
            "single_configs": single_configs,
            "down_configs": down_configs,
            "right_configs": right_configs,
        }
    return guide_spec


def get_data_from_form_with_guide_spec(guide_spec, in_filename, date_format=None):
    data = {}
    in_workbook = openpyxl.load_workbook(in_filename, read_only=True)

    for worksheet_title, worksheet_spec in guide_spec["worksheets"].items():

        # Step 1: Process single configs (easy ones)
        for single_config in worksheet_spec["single_configs"].values():
            json_set_deep_value(
                data,
                single_config["path"],
                _get_cell_value(
                    in_workbook[worksheet_title][single_config["coordinate"]],
                    date_format,
                ),
            )

        # Step 2: Process Down Configs
        for down_config in worksheet_spec["down_configs"].values():
            start_row = down_config[0]["row"]
            max_row = in_workbook[worksheet_title].max_row + 1
            json_set_deep_value(data, down_config[0]["list_path"], [])
            for row in range(start_row, max_row + 1):
                item = {}
                found_anything = False
                for this_down_config in down_config:
                    cell = in_workbook[worksheet_title][
                        this_down_config["column_letter"] + str(row)
                    ]
                    json_set_deep_value(
                        item,
                        this_down_config["item_path"],
                        _get_cell_value(cell, date_format),
                    )
                    if json_get_deep_value(item, this_down_config["item_path"]):
                        found_anything = True
                if found_anything:
                    json_append_deep_value(data, down_config[0]["list_path"], item)

        # Step 3: Process Right Configs
        for right_config in worksheet_spec["right_configs"].values():
            start_column = right_config[0]["column"]
            max_column = in_workbook[worksheet_title].max_column + 1
            json_set_deep_value(data, right_config[0]["list_path"], [])
            for column in range(start_column, max_column + 1):
                item = {}
                found_anything = False
                for this_right_config in right_config:
                    cell = in_workbook[worksheet_title][
                        openpyxl.utils.get_column_letter(column)
                        + str(this_right_config["row"])
                    ]
                    json_set_deep_value(
                        item,
                        this_right_config["item_path"],
                        _get_cell_value(cell, date_format),
                    )
                    if json_get_deep_value(item, this_right_config["item_path"]):
                        found_anything = True
                if found_anything:
                    json_append_deep_value(data, right_config[0]["list_path"], item)

    return data


def put_data_in_form(guide_filename, data, out_filename):
    copyfile(guide_filename, out_filename)
    workbook = openpyxl.load_workbook(out_filename)

    for worksheet in workbook.worksheets:

        # Step 1: build details of all configs on this sheet
        single_configs, down_configs, right_configs = _build_all_configs_in_excel_sheet(
            worksheet
        )

        # Step 2: Process single configs (easy ones)
        for single_config in single_configs.values():
            worksheet[single_config["coordinate"]] = json_get_deep_value(
                data, single_config["path"]
            )

        # Step 3: Process Down Configs
        for down_config in down_configs.values():
            datas_to_insert = json_get_deep_value(data, down_config[0]["list_path"])
            if isinstance(datas_to_insert, list) and len(datas_to_insert) > 0:
                extra_row = 0
                for data_to_insert in datas_to_insert:
                    for this_down_config in down_config:
                        worksheet[
                            this_down_config["column_letter"]
                            + str(this_down_config["row"] + extra_row)
                        ] = json_get_deep_value(
                            data_to_insert, this_down_config["item_path"]
                        )
                    extra_row += 1
            else:
                # no data, but we still want to remove the special values from the output spreadsheet
                for this_down_config in down_config:
                    worksheet[this_down_config["coordinate"]] = ""

        # Step 4: Process Right Configs
        for right_config in right_configs.values():
            datas_to_insert = json_get_deep_value(data, right_config[0]["list_path"])
            if isinstance(datas_to_insert, list) and len(datas_to_insert) > 0:
                extra_column = 0
                for data_to_insert in datas_to_insert:
                    for this_right_config in right_config:
                        worksheet[
                            openpyxl.utils.get_column_letter(
                                this_right_config["column"] + extra_column
                            )
                            + str(this_right_config["row"])
                        ] = json_get_deep_value(
                            data_to_insert, this_right_config["item_path"]
                        )
                    extra_column += 1
            else:
                # no data, but we still want to remove the special values from the output spreadsheet
                for this_right_config in right_config:
                    worksheet[this_right_config["coordinate"]] = ""

    workbook.save(out_filename)
