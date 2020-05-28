import os

import openpyxl

from spreadsheetforms.api import put_data_in_form

TEST_DATA_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data")
TEST_DATA_OUT_DIR = os.path.join(
    os.path.dirname(os.path.abspath(__file__)), "data", "out"
)


def test_1():

    outfile = os.path.join(TEST_DATA_OUT_DIR, "put_data_in_form_1.xlsx")

    data = {
        "noise": "Woof Woof",
        "pet": "Dog",
        "hungry": [
            {"state": "Yes", "wants": "Food"},
            {"state": "Always", "wants": "Food"},
            {"state": "Right Now", "wants": "Food"},
        ],
        "sleepy": [
            {"state": "A lot", "wants": "Sleep"},
            {"state": "Also Right Now", "wants": "A Nap"},
        ],
        "toys": [
            {"title": "Plastic bone", "squeak": "Oh Yes"},
            {"title": "Tennis Ball", "squeak": "No"},
        ],
    }

    put_data_in_form(os.path.join(TEST_DATA_DIR, "pet1.xlsx"), data, outfile)

    workbook = openpyxl.load_workbook(outfile, read_only=True)

    # Test - General Info
    assert "Dog" == workbook["Info"]["B5"].value
    assert "Woof Woof" == workbook["Info"]["B6"].value

    # Test - Hungry
    assert "Yes" == workbook["Info"]["B9"].value
    assert "Food" == workbook["Info"]["B10"].value
    assert "Always" == workbook["Info"]["C9"].value
    assert "Food" == workbook["Info"]["C10"].value
    assert "Right Now" == workbook["Info"]["D9"].value
    assert "Food" == workbook["Info"]["D10"].value

    # Test - Sleepy
    assert "A lot" == workbook["Info"]["B13"].value
    assert "Sleep" == workbook["Info"]["B14"].value
    assert "Also Right Now" == workbook["Info"]["C13"].value
    assert "A Nap" == workbook["Info"]["C14"].value

    # Test - Toys
    assert "Plastic bone" == workbook["Toys"]["A7"].value
    assert "Oh Yes" == workbook["Toys"]["B7"].value
    assert "Tennis Ball" == workbook["Toys"]["A8"].value
    assert "No" == workbook["Toys"]["B8"].value


def test_no_list():

    outfile = os.path.join(TEST_DATA_OUT_DIR, "put_data_in_form_1_no_list.xlsx")

    data = {"noise": "Woof Woof", "pet": "Dog"}

    put_data_in_form(os.path.join(TEST_DATA_DIR, "pet1.xlsx"), data, outfile)

    workbook = openpyxl.load_workbook(outfile, read_only=True)

    # Test - General Info
    assert "Dog" == workbook["Info"]["B5"].value
    assert "Woof Woof" == workbook["Info"]["B6"].value

    # Test - Hungry
    assert None == workbook["Info"]["B9"].value
    assert None == workbook["Info"]["B10"].value
    assert None == workbook["Info"]["C9"].value
    assert None == workbook["Info"]["C10"].value
    assert None == workbook["Info"]["D9"].value
    assert None == workbook["Info"]["D10"].value

    # Test - Sleepy
    assert None == workbook["Info"]["B13"].value
    assert None == workbook["Info"]["B14"].value
    assert None == workbook["Info"]["C13"].value
    assert None == workbook["Info"]["C14"].value

    # Test - Toys
    assert None == workbook["Toys"]["A7"].value
    assert None == workbook["Toys"]["B7"].value
    assert None == workbook["Toys"]["A8"].value
    assert None == workbook["Toys"]["B8"].value


def test_deep():

    outfile = os.path.join(TEST_DATA_OUT_DIR, "put_data_in_form_1_deep.xlsx")

    data = {
        "emits": {"noise": "Woof Woof"},
        "pet": {"kind": "Dog"},
        "mood": {
            "hungry": [
                {"current": {"state": "Yes", "wants": "Food"}},
                {"current": {"state": "Always", "wants": "Food"}},
                {"current": {"state": "Right Now", "wants": "Food"}},
            ],
            "sleepy": [
                {"current": {"state": "A lot", "wants": "Sleep"}},
                {"current": {"state": "Also Right Now", "wants": "A Nap"}},
            ],
        },
        "likes": {
            "toys": [
                {
                    "human-concerns": {"title": "Plastic bone"},
                    "pet-concerns": {"squeak": "Oh Yes"},
                },
                {
                    "human-concerns": {"title": "Tennis Ball"},
                    "pet-concerns": {"squeak": "No"},
                },
            ]
        },
    }

    put_data_in_form(os.path.join(TEST_DATA_DIR, "pet1-deep.xlsx"), data, outfile)

    workbook = openpyxl.load_workbook(outfile, read_only=True)

    # Test - General Info
    assert "Dog" == workbook["Info"]["B5"].value
    assert "Woof Woof" == workbook["Info"]["B6"].value

    # Test - Hungry
    assert "Yes" == workbook["Info"]["B9"].value
    assert "Food" == workbook["Info"]["B10"].value
    assert "Always" == workbook["Info"]["C9"].value
    assert "Food" == workbook["Info"]["C10"].value
    assert "Right Now" == workbook["Info"]["D9"].value
    assert "Food" == workbook["Info"]["D10"].value

    # Test - Sleepy
    assert "A lot" == workbook["Info"]["B13"].value
    assert "Sleep" == workbook["Info"]["B14"].value
    assert "Also Right Now" == workbook["Info"]["C13"].value
    assert "A Nap" == workbook["Info"]["C14"].value

    # Test - Toys
    assert "Plastic bone" == workbook["Toys"]["A7"].value
    assert "Oh Yes" == workbook["Toys"]["B7"].value
    assert "Tennis Ball" == workbook["Toys"]["A8"].value
    assert "No" == workbook["Toys"]["B8"].value
